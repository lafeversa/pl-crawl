from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import openpyxl as pxl
import random
import time
from collections import OrderedDict
from os.path import exists
from os import replace
import pickle
import logging as log

'''This script takes a call list for candidate profiles and uses that to lookup
additional information for each of the candidates within that list.
'''

def get_candidates_from_file(driver, candidate_file='C:/Users/lafeversa/Documents/Python/Base%20Jumper/Call%20List%20HM%20OR%20Rural.html'):
    '''Read the call list file to generate a list of candidates.
    
    Returns:
        A list of valid links that correspond to the webpages with additional 
        information for the candidates in `candidate_file`.
    '''
    drivergot = driver.get('file:///' + candidate_file)

    agent = driver.execute_script("return navigator.userAgent")
    print (agent)

    assert 'Manage Call List' in driver.title

    md_links = driver.find_elements_by_partial_link_text('MD')
    do_links = driver.find_elements_by_partial_link_text('DO')

    valid_links = OrderedDict()

    for link in md_links:
        link = link.get_attribute('href')
        valid_links[link] = None
    for link in do_links:
        link = link.get_attribute('href')
        valid_links[link] = None
    return valid_links

def get_all_candidate_data(driver, links, start, stop):
    '''Iterate through all links in `links` and obtain the candidate data from 
    them.
    
    Args:
        links (list): List of the valid links to grab data from.
        start (int): Starting index for the list(links.keys())[start:stop]
        
    Returns:
        all_candidate_data (dict): Dictionary of candidate data. The first key 
            is the index of the candidate from the `link_index` variable. 
            
            So `candidate_values[0]` is a dictionary of the candidate data for 
            the very first candidate that was read from the candidate_file. This 
            dictionary has keys corresponding to the different pieces of 
            candidate information, like 'First', 'Last', 'Specialty', etc.
    '''
    log.debug('Inside get_all_candidate_data.')
    log.debug('Type(links): {}'.format(type(links)))
    log.debug('start: {}, stop: {}'.format(start, stop))
    
    # Iterating through each link we have stored
    #for link in links.keys():
    #print (list(links.keys())[0:10])
    for link in list(links.keys())[start:stop]:
        try:
            driver.get(link)
        except:
            log.error('Failed inside of the get_all_candidate_data function.')
            log.error('\tLink: {}'.format(link))
            log.error('\tType(link): {}'.format(type(link)))
            log.error('\tLink_index: {}'.format(link_index))
            log.error('\tNum keys in all_candidate_data: {}'.format(len(all_candidate_data.keys())))
        wait_time = random.randint(3, 12)
        print ('Security delay is {} seconds.'.format(wait_time))
        
        # If we need to login, do so
        if 'Log In' in driver.title:
            assert 'Log In' in driver.title
            do_login(driver)
            
        if 'Candidate Detail' in driver.title:
            assert 'Candidate Detail' in driver.title
            time.sleep(wait_time)
            
            candidate_data = get_candidate_data(driver, link) 
            links[link] = candidate_data
        else:
            continue
    return links
        
def get_candidate_data(driver, link):
    '''Grab data for a candidate from the webpage at `link`.
    
    Args:
        link (str): URL to the page with additional candidate data
        
    Returns:
        candidate_data (dict): Dictionary of candidate data, where the keys 
            are the various pieces of candidate information, like 'First',
            'Last', 'Specialty', etc.
    '''
    nameofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblCandidateNameValue')
    nameofcand_text = nameofcand.text
    
    print ('Pulling data on: \n', nameofcand_text + '\n')
    
    # Split the name into first, last, degree 
    nameofcand_text = nameofcand_text.split(' ')
    firstnameofcand_text = nameofcand_text[0].strip()
    lastnameofcand_text = nameofcand_text[1].strip(',')
    degreeofcand_text = nameofcand_text[2].strip()
    if degreeofcand_text == 'MD':
        degreeofcand_text = 'M.D.'
    elif degreeofcand_text == 'DO':
        degreeofcand_text = 'D.O.'
    else:
        pass

    specialtyofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblProfessionSpecialty')
    specialtyofcand_text = specialtyofcand.text
    
    
    if 'Physician - ' in specialtyofcand_text:
        specialtyofcand_text = specialtyofcand_text.strip('Physician - ')
    
    if 'Emergency' in specialtyofcand_text and 'Pediatrics' not in specialtyofcand_text:
        specialtyofcand_text = 'Emergency Medicine'
    elif 'Hospitalist' in specialtyofcand_text:
        specialtyofcand_text = 'Hospital Medicine'
    elif 'Anesthesiology' in specialtyofcand_text:
        specialtyofcand_text = 'Anesthesiology'
    elif 'Intensivist' in specialtyofcand_text:
        specialtyofcand_text = 'Intensivist'
    elif 'Urgent' in specialtyofcand_text:
        specialtyofcand_text = 'Urgent Care'
    elif 'Pediatrics' in specialtyofcand_text and 'Emergency' in specialtyofcand_text:
        specialtyofcand_text = 'Pediatric Emergency Medicine'
    else:
        pass
    
    homephoneofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblHomePhoneValue')
    homephoneofcand_text = homephoneofcand.text
    
    cellphoneofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblCellPhoneValue')
    cellphoneofcand_text = cellphoneofcand.text
    
    emailofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_HypEmailValue')
    emailofcand_text = emailofcand.text
    
    geoprefofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidatePersonalInformationSummary_LblGeographicPreferenceValue')
    geoprefofcand_text = geoprefofcand.text
    
    detailedgeoprefofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidatePersonalInformationSummary_LblDetailGeoPreferenceValue')
    detailedgeoprefofcand_text = detailedgeoprefofcand.text
    
    addressofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblCandidateAddressValue')
    addressofcand_text = addressofcand.text
    
    citystateofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateContactInfo1_LblLocationValue')
    citystateofcand_text = citystateofcand.text
    
    # Splitting the city, state, zip info into their own fields
    citystateofcand_text = citystateofcand_text.split(' ')
    if len(citystateofcand_text) == 3:
        cityofcand_text = citystateofcand_text[0].strip(',')
        stateofcand_text = citystateofcand_text[1].strip()
        zipofcand_text = citystateofcand_text[2].strip()
    elif len(citystateofcand_text) == 4:
        cityofcand_text = citystateofcand_text[0].strip(',') + ' ' + citystateofcand_text[1].strip(',')
        stateofcand_text = citystateofcand_text[2].strip()
        zipofcand_text = citystateofcand_text[3].strip()
    elif len(citystateofcand_text) == 5:
        cityofcand_text = citystateofcand_text[0].strip(',') + ' ' + citystateofcand_text[1].strip(',') + ' ' + citystateofcand_text[2].strip(',')
        stateofcand_text = citystateofcand_text[3].strip()
        zipofcand_text = citystateofcand_text[4].strip()
    else:
        cityofcand_text = 'Not provided'
        stateofcand_text = 'Not provided'
        zipofcand_text = 'Not provided'
    
    licenseofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateProfessionalStatusSummary_LblLicensedValue')
    licenseofcand_text = licenseofcand.text
    
    residencyofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateProfessionalStatusSummary_LblResidencyValue')
    residencyofcand_text = residencyofcand.text
    
    citizenshipofcand = driver.find_element_by_id('ContentBody_ContentBody_ContentBody_CandidateProfileTabs1_CusCandidateProfessionalStatusSummary_LblCitizenshipStatusValue')
    citizenshipofcand_text = citizenshipofcand.text
    
    descriptionofcand_text = 'Found on PracticeLink - Resume may be there'

    candidate_values = {'First': str(firstnameofcand_text),
                        'Last': str(lastnameofcand_text),
                        'Title': str(degreeofcand_text),
                        'Specialty': str(specialtyofcand_text),
                        'Home Phone': str(homephoneofcand_text), 
                        'Cell Phone': str(cellphoneofcand_text),
                        'Email': str(emailofcand_text),
                        'Geo Pref': str(geoprefofcand_text),
                        'Detailed Pref': str(detailedgeoprefofcand_text),
                        'Address': str(addressofcand_text), 
                        'City': str(cityofcand_text),
                        'State': str(stateofcand_text),
                        'Zip': str(zipofcand_text),
                        'License(s)': str(licenseofcand_text),
                        'Residency': str(residencyofcand_text), 
                        'Citizenship Status': str(citizenshipofcand_text),
                        'Description': str(descriptionofcand_text)
    }

    return candidate_values
           
def do_login(driver):
    username = driver.find_element_by_id('ctl00_ctl00_ContentBody_ContentBody_LoginControl_TxtEmail')
    username.send_keys('username')
    
    password = driver.find_element_by_id('ctl00_ctl00_ContentBody_ContentBody_LoginControl_PwdPassword')
    password.send_keys('password')
    
    loginbutton = driver.find_element_by_id('ctl00_ctl00_ContentBody_ContentBody_LoginControl_BtnLogIn')
    loginbutton.send_keys(Keys.RETURN)

def initialize_candidate_list(driver, candidate_list_saved='candidate_list.pk1'):
    '''
    Look for a data file with pickled dictionary information and load it 
    if it exists. If no such file, load the default dictionaries.
    '''
    
    if exists(candidate_list_saved):
        cand_list_file = open(candidate_list_saved, 'rb')
        # Load in the tuple of (keys, values)
        candidate_links = pickle.load(cand_list_file)
        cand_list_file.close()
        
        log.debug('Loading candidate data from pickle file.')
        log.debug('Type of candidate_links: {}'.format(type(candidate_links)))
        if type(candidate_links) == type((1,2)):
            log.debug('\tType of 0th element: {}'.format(type(candidate_links[0])))
            log.debug('\tType of 1st element: {}'.format(type(candidate_links[1])))
        
        # Create a new OrderedDict and populate it with the saved data we just read
        pickled_keys = candidate_links[0]
        pickled_vals = candidate_links[1]
        candidate_links_rebuilt = OrderedDict()
        for p in list(zip(pickled_keys, pickled_vals)):
            k = p[0]
            v = p[1]
            candidate_links_rebuilt[k] = v
        candidate_links = candidate_links_rebuilt
    else:
        log.debug('Loading candidate data from saved webpage.')
        candidate_links = get_candidates_from_file(driver)
        
    log.debug('{} keys, {} values. {} 0th key, {} last key.'.format(
        len(candidate_links.keys()), len(candidate_links.values()), 
        list(candidate_links.keys())[0], list(candidate_links.keys())[-1]
        ))
        
    return candidate_links

def get_last_read_candidate(candidate_links):
    '''Magic
    '''
    prev_key = None
    prev_val = 'is not truly needed'
    
    for k, v in candidate_links.items():
        if prev_key is None:
            prev_key = k
            prev_val = v
            continue
            
        if v is None and prev_val is None:
            return prev_key
        else:
            prev_key = k
            prev_val = v
    return prev_key
    
def write_output(candidates, savename='jumpedoutput.xlsx'):
    '''Write the candidate data to an Excel workbook. The header values of the 
    workbook can be adjusted by changing the `headers` list in this function 
    and the candidate data will follow the same order.     
    '''
    newWB = pxl.Workbook()
    sheet = newWB.active
    
    headers = ['First', 'Last', 'Title', 'Specialty', 'Home Phone', 
                'Cell Phone', 'Email', 'Geo Pref', 'Detailed Pref', 'Address', 
                'City', 'State', 'Zip', 'License(s)', 'Residency', 
                'Citizenship Status', 'Description']
    
    for h in range(len(headers)):
        sheet.cell(row = 1, column = h + 1).value = headers[h]
    
    # For each person, write the person's information that corresponds to each 
    # header item.
    for person in candidates.keys():
        row_index = list(candidates.keys()).index(person) + 2
        for h in headers:
            col_index = headers.index(h) + 1
            if candidates[person] is not None:
                sheet.cell(row = row_index, column = col_index).value = candidates[person][h]
            
    newWB.save(savename)

def save_candidate_data(candidate_data, candidate_list_saved='candidate_list.pk1'):
    '''We save the candidate data as a tuple of the OrderedDict's (keys,values). 
    '''
    # Optional, but if there is saved data with this filename already present
    # this will rename the current save file by adding '.bak' to the end. This 
    # will overwrite any other files of the same name (previous day's saves). 
    # So just the previous day's save point is kept as backup.
    log.debug('Saving candidate data.')
    log.debug('{} keys, {} values. {} 0th key, {} last key.'.format(
        len(candidate_data.keys()), len(candidate_data.values()), 
        list(candidate_data.keys())[0], list(candidate_data.keys())[-1]
        ))
        
    if exists(candidate_list_saved):
        replace(candidate_list_saved, candidate_list_saved + '.bak')
        
    print('Saving candidate data to {}'.format(candidate_list_saved))
    with open(candidate_list_saved, 'wb') as fout:
        pickle.dump((list(candidate_data.keys()), list(candidate_data.values())), fout)

def test_get_last_read():
    d = test_data()
    assert get_last_read_candidate(d) == 'sixth_key'
        
def test_data():
    d = OrderedDict()
    d['first_key'] = test_candidate_dict(1)
    d['second_key'] = test_candidate_dict(2)
    d['third_key'] = test_candidate_dict(3)
    d['fourth_key'] = None
    d['fifth_key'] = test_candidate_dict(5)
    d['sixth_key'] = None
    d['seventh_key'] = None
    return d
    
def test_candidate_dict(i):
    d = {'First': i,
        'Last': i,
        'Title': i,
        'Specialty': i,
        'Home Phone': i, 
        'Cell Phone': i,
        'Email': i,
        'Geo Pref': i,
        'Detailed Pref': i,
        'Address': i, 
        'City': i,
        'State': i,
        'Zip': i,
        'License(s)': i,
        'Residency': i, 
        'Citizenship Status': i,
        'Description': i
    }
    return d


def get_starting_indexes(candidate_links, place_to_start):
    if place_to_start is None:
        start_index = 0
    else:
        start_index = list(candidate_links.keys()).index(place_to_start)
    if start_index == len(candidate_links.keys()):
        print ('Everything has been read.')
        return None, None
    stop_index = start_index + daily_reads
    return start_index, stop_index

def run():
    log.basicConfig(filename='basejumper_log.txt', level=log.DEBUG)
    driver = webdriver.Firefox()

    savename = 'jumpedoutput.xlsx'
    daily_reads = random.randint(6, 8)
    print ('Reading {} candidates this run.'.format(daily_reads))
    
    candidate_links = initialize_candidate_list(driver)

    place_to_start = get_last_read_candidate(candidate_links)
    start_index, stop_index = get_starting_indexes(candidate_links, place_to_start)
    
    if start_index is None and stop_index is None:
        return
    candidates = get_all_candidate_data(driver, candidate_links, start_index, stop_index)    

    write_output(candidates, savename)
    print('Done reading candidate information for candidates ' + 
          '{} through {} of {}.'.format(start_index, stop_index,
                                len(candidates.keys())))
                                
    save_candidate_data(candidates)
    
if __name__ == '__main__':
    run()
    

    

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
